import requests
import json
from authentication import *
from utils import *
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
import warnings
from dotenv import load_dotenv

warnings.filterwarnings("ignore")
load_dotenv()



#Obtener los Feature IDs sólo para los Feature Templates con nombre VPN0 y tipo vedge interface
def get_all_feat_template(header,base_url):
    dictaux={}
    listaux=[]
    
    feat_temp_url="dataservice/template/feature"

    print("API CALL...")
    all_feat_response=requests.get(url=f"{base_url}{feat_temp_url}",headers=header,verify=False)

    if all_feat_response.status_code==200:
        print("API request was succesful", all_feat_response.status_code)
        all_feat_responsedata=all_feat_response.json()["data"]
        #print(json.dumps(all_feat_response.json(),indent=4))
        for dictio in all_feat_responsedata:
            #mfil=re.search(r'*Gre',dictio["templateName"])
            ##if dictio["templateType"]=="vpn-vedge-interface-gre":
                print(f"Feature Template {dictio['templateName']} found, saving data...")
            
                listaux.append({"FeatureTemplateName":dictio['templateName'],"FeatureTemplateDescription":dictio["templateDescription"],
                "FeatureTemplateId":dictio['templateId'],"TemplateType":dictio['templateType']})
    
        dictaux["data"]=listaux
        savefile("all_feat_temp.json",dictaux)

    else:
        print("Error",all_feat_response.text)

#Obtener los Device Templates IDs dado un Feature Template ID
def get_device_feat_temp(header,base_url):
    dictdevfeat = {}
    listdevcetemp=[]

    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}
    
    with open("all_feat_temp.json",'r') as jaf:
        jafile=jaf.read()
        jparse=json.loads(jafile)["data"]
    for dix in jparse:
        print(f"Getting Device Templates associated with {dix['FeatureTemplateName']}...")
        devtempurl=f"dataservice/template/feature/devicetemplates/{dix['FeatureTemplateId']}"
        devicetempresp=requests.get(url=f"{base_url}{devtempurl}",headers=header,verify=False)
        if devicetempresp.status_code==200:
            devicetempresp=devicetempresp.json()
            #print(json.dumps(devicetempresp,indent=4))
            listdevcetemp.append({"FeatureTemplateName":dix["FeatureTemplateName"],"FeaturetemplateId":dix["FeatureTemplateId"],"FTType":dix["TemplateType"],"DeviceTemplates":devicetempresp["data"]})
           
        else:
            print("Error retrievieng data...",devicetempresp.status_code)

    dictdevfeat["data"]=listdevcetemp
    savefile("all_device-feat_templates.json",dictdevfeat)

#Obtener los Devices que están asociados a un Device Template
def get_devices(header,base_url):
    dictdev = {}
    listdev=[]
    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}
    
    with open("all_device-feat_templates.json",'r') as jaf:
        jafile=jaf.read()
        jparse=json.loads(jafile)["data"]

    for dix in jparse:
        for devtemp in dix["DeviceTemplates"]:
            devurl=f"dataservice/template/device/config/attached/{devtemp['templateId']}"
            print(f"Searching for devices attached to {devtemp['templateName']}")
            response=requests.get(url=f"{base_url}{devurl}",headers=header,verify=False)
            if response.status_code==200:
                
                devresponse=response.json()["data"]
                for device in devresponse:
                    listdev.append({"FeatureTemplateName":dix["FeatureTemplateName"],"FeatureTemplateType":dix["FTType"],"DeviceTemplate":devtemp["templateName"],"Hostname":device["host-name"]})

            else:
                print("Error retrievieng data...",response.status_code)

    dictdev["data"]=listdev
    savefile("all_devices_templates_.json",dictdev)

#Obtener los PolicyIDs 

def save_to_excel(excelfile):

    # workbook = load_workbook(filename=excelfile)
    # sheet = workbook.create_sheet("device-feature")

    book = Workbook()
    sheet = book.active

    with open("all_device-feat_templates.json",'r') as f:
        jparse=f.read()
        jfile=json.loads(jparse)["data"]

    sheet.cell(row=1, column=1).value = "Feature Template name"
    sheet.cell(row=1, column=2).value = "Feature Template type"
    sheet.cell(row=1, column=3).value = "Device Template"
    
    rows=2
    for data in jfile:
        #sheet.cell(row=rows, column=1).value = data["FeatureTemplateName"]
        for devtemp in data["DeviceTemplates"]:
            sheet.cell(row=rows, column=1).value = data["FeatureTemplateName"]
            sheet.cell(row=rows, column=2).value = devtemp["templateName"]
            rows+=1
            
    print("Saving spreadsheet")
    #workbook.save(excelfile)
    book.save(excelfile)

def save_devices(excelfile):
    book = Workbook()
    sheet = book.active
    with open("all_devices_templates_.json",'r') as f:
        jparse=f.read()
        jfile=json.loads(jparse)["data"]
    
    sheet.cell(row=1, column=1).value = "Feature Template name"
    sheet.cell(row=1, column=2).value = "Device Template type"
    sheet.cell(row=1, column=3).value = "Device Template"
    sheet.cell(row=1, column=4).value = "Device"

    rows=2
    for data in jfile:
        sheet.cell(row=rows, column=1).value = data["FeatureTemplateName"]
        sheet.cell(row=rows, column=2).value = data["FeatureTemplateType"]
        sheet.cell(row=rows, column=3).value = data["DeviceTemplate"]
        sheet.cell(row=rows, column=4).value = data["Hostname"]
        rows+=1
    print("Saving spreadsheet")
    book.save(excelfile)



#Set header
def setheader(jsessionid,token):
    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}

    return header




if __name__ == "__main__":

    base_url=os.getenv('SDWAN_URL')
    today = date.today()
    excelfile=f"Device-device-feat-template_{today}.xlsx"
    username=os.getenv('SDWAN_USERNAME')
    password=os.getenv('SDWAN_PASSWORD')
    jsessionid=get_jsessionid(base_url,username,password)
    token=get_token(jsessionid,base_url)
    header=setheader(jsessionid,token)
    get_all_feat_template(header,base_url)
    get_device_feat_temp(header,base_url)
    get_devices(header,base_url)
    #save_to_excel(excelfile)
    save_devices(excelfile)



import requests
import json
import getpass
import os.path
from authentication import *
from utils import *
from html_python import *
from time import time
from time import sleep
from sys import stdout as terminal
from itertools import cycle
from threading import Thread
from openpyxl import Workbook

__author__ = "Mario Uriel Romero Martinez"
__version__ = "1.0"
__maintainer__ = "Mario Uriel Romero Martinez"


def animate():
    for c in cycle(['|', '/', '-', '\\']):
        if done:
            break
        terminal.write('\rworking ' + c)
        terminal.flush()
        sleep(0.1)
    terminal.write('\rDone!    ')
    terminal.flush()


#Set header
def setheader(jsessionid,token):
    if token is not None:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid, 'X-XSRF-TOKEN': token}
    else:
        header = {'Content-Type': "application/json", 'Cookie': jsessionid}

    return header

# Get All Localized Policies
def getlp(header,base_url,lpjfile):
    listlp=[]
    dictlp={}

    lpurl="dataservice/template/policy/vedge"
    try:
        print("Getting Localized Policy information")
        lpresponse=requests.get(url=f"{base_url}{lpurl}",headers=header,verify=False)
        lpresponse.raise_for_status()
        if lpresponse.status_code==200:
            lpdata=lpresponse.json()["data"]

            for lp in lpdata:
                listlp.append({"lpname":lp['policyName'],"lpid":lp["policyId"]})
   
    
        else:
            print("Error", lpresponse.status_code,lpresponse.text)

    except requests.exceptions.HTTPError as httpe:
        print("HTTP Request error",httpe)
        sys.exit()
    except requests.exceptions.Timeout as toe:
        print("HTTP Request error",toe)
        sys.exit()
    except requests.exceptions.ConnectionError as coe:
        print("HTTP Request error",coe)
        sys.exit()
    except requests.exceptions.RequestException as err:
        print("HTTP Request error",err)
        sys.exit()



    dictlp["data"]=listlp
    savefile(lpjfile,dictlp)   


#Get all Device templates
def getdt(header,base_url,jfile):

    dictdevfeat = {}
    listdevcetemp=[]

    global done
    done= False

    print("Getting Device Template Information")
    t = Thread(target=animate)
    t.start()

    dturl="dataservice/template/device"
    try:
        dtresponse=requests.get(url=f"{base_url}{dturl}",headers=header,verify=False)
        dtresponse.raise_for_status()
        if dtresponse.status_code==200:
            dtdata=dtresponse.json()["data"]
            for dtd in dtdata:
                listdevcetemp.append({"dtname":dtd["templateName"],"dtid":dtd["templateId"]})

        else:
            print("Error", dtresponse.status_code,dtresponse.text)

    except requests.exceptions.HTTPError as httpe:
        print("HTTP Request error",httpe)
        done = True
        sys.exit()
    except requests.exceptions.Timeout as toe:
        print("HTTP Request error",toe)
        done = True
        sys.exit()
    except requests.exceptions.ConnectionError as coe:
        print("HTTP Request error",coe)
        done = True
        sys.exit()
    except requests.exceptions.RequestException as err:
        print("HTTP Request error",err)
        done = True
        sys.exit()

    dictdevfeat["data"]=listdevcetemp
    savefile(jfile,dictdevfeat)
    done = True
 

#Get attached Localized Policy Id from a given Device template Id
def getlpfromdt(header,base_url,dtjfile,dtlpjfile):
    listdevtemp=[]
    dictdev= {}

    with open(dtjfile,'r') as jaf:
        jafile=jaf.read()
        jparse=json.loads(jafile)["data"]
    
    global done
    done= False

    print("Getting associated Localized Policy to a Device Template")

    t = Thread(target=animate)
    t.start()

    for devtempid in jparse:

        fturl=f"dataservice/template/device/object/{devtempid['dtid']}"
        try:
            ftresponse=requests.get(url=f"{base_url}{fturl}",headers=header,verify=False)
            ftresponse.raise_for_status()
            if ftresponse.status_code==200:
                ftdata=ftresponse.json()
                #print(json.dumps(ftdata, indent=4))
                if "policyId" in ftdata:
                    #devtempid["policyId"]=ftdata["policyId"]
                    pid=ftdata["policyId"]
                    #print(ftdata["policyId"])
                else:
                    #devtempid["policyId"]=None
                    pid=None
                listdevtemp.append({"devtempid":devtempid['dtid'],"devtemppolid":pid})
   
            else:
                print("Error", ftresponse.status_code,ftresponse.text)

        except requests.exceptions.HTTPError as httpe:
            print("HTTP Request error",httpe)
            done = True
            sys.exit()
            
        except requests.exceptions.Timeout as toe:
            print("HTTP Request error",toe)
            done = True
            sys.exit()
           
        except requests.exceptions.ConnectionError as coe:
            print("HTTP Request error",coe)   
            done = True
            sys.exit()
        except requests.exceptions.RequestException as err:
            print("HTTP Request error",err)
            done = True
            sys.exit()
            

    dictdev["data"]=listdevtemp
    savefile(dtlpjfile,dictdev)
    done = True


#Find the relationship between LPId and Device Template Id
def create_relationship(lpjfile,dtjfile,dtlpjfile,header,base_url):
    
    dictaux={}
  
    with open(lpjfile,'r') as lpjaf:
        lpjafile=lpjaf.read()
        lpjparse=json.loads(lpjafile)["data"]
    

    with open(dtlpjfile,'r') as dtlpjaf:
        dtlpjafile=dtlpjaf.read()
        dtlpjparse=json.loads(dtlpjafile)["data"]

    global done
    done= False
    print("Matching relationship between a Localized Policy and Device Template")
    t = Thread(target=animate)
    t.start()

    for lp in lpjparse:
        listdt=[]
        for dtlp in dtlpjparse:
            if lp["lpid"]==dtlp["devtemppolid"]:
                listdt.append(dtlp["devtempid"])

        listdtname=search(listdt,header,base_url,dtjfile)
        lp["devtemp"]=listdtname
        
    dictaux["data"]=lpjparse
    savefile(lpjfile,dictaux)
    done = True

#Get devices given a Device TemplateId
def search(listdtid,header,base_url,dtjfile):
    listname=[]

    with open(dtjfile,'r') as dtjaf:
        dtjafile=dtjaf.read()
        dtjparse=json.loads(dtjafile)["data"]

    global done
    done= False


    print("Getting Devices info")
    t = Thread(target=animate)
    t.start()

    for dtid in listdtid:
        devlist=[]
        devurl=f"dataservice/template/device/config/attached/{dtid}"
        try:
            devresponse=requests.get(url=f"{base_url}{devurl}",headers=header,verify=False)
            devresponse.raise_for_status()
            if devresponse.status_code==200:
                devdata=devresponse.json()["data"]
                for device in devdata:
                    if "site-id" in device:
                        siteid=device["site-id"]
                    else:
                        siteid=None

                    devlist.append({"dname":device["host-name"],"systemip":device["deviceIP"],"siteid":siteid})
            else:
                print("Error", devresponse.status_code,devresponse.text)

        except requests.exceptions.HTTPError as httpe:
            print("HTTP Request error",httpe)
            done = True
            sys.exit()
           
        except requests.exceptions.Timeout as toe:
            print("HTTP Request error",toe)
            done = True
            sys.exit()
            
        except requests.exceptions.ConnectionError as coe:
            print("HTTP Request error",coe)  
            done = True
            sys.exit()

        except requests.exceptions.RequestException as err:
            print("HTTP Request error",err)
            done = True 
            sys.exit()
              

        for dt in dtjparse:
            if dtid==dt["dtid"]:
                listname.append({dt["dtname"]:devlist})
    
    done = True
    return listname
    

#Get LP definition-> Localized policy types
def get_lp_types(header,base_url,lpjfile):

    dictaux={}

    with open(lpjfile,'r') as lpjaf:
        lpjafile=lpjaf.read()
        lpjparse=json.loads(lpjafile)["data"]
    
    global done
    done= False
    print("Getting Localized Policy definitions")
    t = Thread(target=animate)
    t.start()
    
    for lp in lpjparse:
        #lptypes=[]
        lpcdurl=f"dataservice/template/policy/vedge/definition/{lp['lpid']}"
        try:
            lpcresponse=requests.get(url=f"{base_url}{lpcdurl}",headers=header,verify=False)
            lpcresponse.raise_for_status()
            if lpcresponse.status_code==200:
                lpcdata=lpcresponse.json()
                if "assembly" in lpcdata["policyDefinition"]:
                #print(lpcdata["policyDefinition"])
                    lp["lptypes"]=lpcdata["policyDefinition"]["assembly"]
            else:
                print("Error", lpcresponse.status_code,lpcresponse.text)

        except requests.exceptions.HTTPError as httpe:
            print("HTTP Request error",httpe)
            done = True
            sys.exit()
           
        except requests.exceptions.Timeout as toe:
            print("HTTP Request error",toe)
            done = True
            sys.exit()
           
        except requests.exceptions.ConnectionError as coe:
            print("HTTP Request error",coe)  
            done = True 
            sys.exit()
        
        except requests.exceptions.RequestException as err:
            print("HTTP Request error",err)
            done = True
            sys.exit()

    dictaux["data"]=lpjparse
    savefile(lpjfile,dictaux)
    done = True


#Get Localized Policy configuration by type: Forwarding Class/QoS, Access Control Lists, Route Policy
def get_lp_config(header,base_url,lpjfile):

    dictaux={}


    with open(lpjfile,'r') as lpjaf:
        lpjafile=lpjaf.read()
        lpjparse=json.loads(lpjafile)["data"]


    global done
    done= False
    print("Getting Localized Control/Data policies")
    t = Thread(target=animate)
    t.start()

    for lp in lpjparse:
        if "lptypes" in lp:
            for lpt in lp["lptypes"]:
                try:
                    lpturl=f"dataservice/template/policy/definition/{lpt['type'].lower()}/{lpt['definitionId']}"
                    lptresponse=requests.get(url=f"{base_url}{lpturl}",headers=header,verify=False)
                    lptresponse.raise_for_status()
                    if lptresponse.status_code==200:
                        lptdata=lptresponse.json()
                        if "sequences" in lptdata:
                            lpt["policyconfig"]={"policyname":lptdata["name"],"sequences":lptdata["sequences"]}
                        
                        if "definition" in lptdata:
                            lpt["policyconfig"]={"policyname":lptdata["name"],"sequences":lptdata["definition"]["qosSchedulers"]}


                    else:
                        print("Error", lptresponse.status_code,lptresponse.text)

                except requests.exceptions.HTTPError as httpe:
                    print("HTTP Request error",httpe)
                    done = True
                    sys.exit()
                    
                except requests.exceptions.Timeout as toe:
                    print("HTTP Request error",toe)
                    done = True
                    sys.exit()
                  
                except requests.exceptions.ConnectionError as coe:
                    print("HTTP Request error",coe)   
                    done = True
                    sys.exit()
                     
                except requests.exceptions.RequestException as err:
                    print("HTTP Request error",err)
                    done = True
                    sys.exit()

    dictaux["data"]=lpjparse
    savefile(lpjfile,dictaux)
    done = True




def buildtreemap(lpjfile,lptreedata):

    path="lptree.json"
    check_file=os.path.isfile(path)
    if check_file:

        with open("lptree.json",'r') as tjaf:
            tjafile=tjaf.read()
            treest=json.loads(tjafile)

        with open(lpjfile,'r') as lpjaf:
            lpjafile=lpjaf.read()
            lpjparse=json.loads(lpjafile)["data"]

        print("Creating treemap file")
        i=0
        for lp in lpjparse:
            listaux=[]
            if "lptypes" in lp:
                for lpt in lp["lptypes"]:
                    if "policyconfig" in lpt:
                        lpnd=lpt["policyconfig"]["policyname"]
                    else:
                        lpnd=None
                    
                    listaux.append({lpt["type"]:lpnd})

                        #treest["children"].append({"name":lp["lpname"],"wcol":"purple","lptype":lpt["type"],"lptname":lpt["policyconfig"]["policyname"],"children":[]})
        
                #treest["children"].append({"name":lp["lpname"],"wcol":"purple","lptypes":listaux,"children":[]})
            #else:
                #treest["children"].append({"name":lp["lpname"],"wcol":"purple","children":[]})
            treest["children"].append({"name":lp["lpname"],"wcol":"purple","lptypes":listaux,"children":[]})
                

            j=0
            for dt in lp["devtemp"]:
                for dtname,device in dt.items():
                    treest["children"][i]["children"].append({"name":dtname,"wcol":"green","children":[]})
                    for dev in device: 
                        treest["children"][i]["children"][j]["children"].append({"name":dev["dname"],"sip":dev["systemip"],"siteid":dev["siteid"],"wcol":"steelblue"})  

                j+=1
            i+=1
        
        savefile(lptreedata,treest)    
    else:
        print("Check if lptree.json exists in the local directory")


def count_dt_dev(treejsonf):

    cntrdtlist=[]
    cntrdlist=[] 

    with open(treejsonf,'r') as tjaf:
        tjafile=tjaf.read()
        treef=json.loads(tjafile)
    

    for lp in treef["children"]:
        dtcounter=0
        for dt in lp["children"]:
            dtcounter+=1
            dcounter=0
            for dev in dt["children"]:
                dcounter+=1
            cntrdlist.append(dcounter)
        cntrdtlist.append(dtcounter)
        

    return cntrdtlist,cntrdlist
  

def logouts(header,base_url):
    lourl=f"{base_url}logout?nocache={str(int(time()))}"
    payload={}

    try:
        response = requests.request("GET", lourl, headers=header, data=payload, verify=False)

        if response.status_code == 200:
            print("Session closed succesfully")
        else:
            print("Unable to close the session",response.status_code)

    except requests.exceptions.HTTPError as httpe:
        print("HTTP Request error",httpe)
        done = True
        sys.exit()
    except requests.exceptions.Timeout as toe:
        print("HTTP Request error",toe)
        done = True
        sys.exit()
    except requests.exceptions.ConnectionError as coe:
        print("HTTP Request error",coe)
        done = True
        sys.exit()
    
    except requests.exceptions.RequestException as err:
        print("HTTP Request error",err)
        done = True
        sys.exit()


def convert_excel(treef,excelf):

    print("Creating spreadsheet")
    book = Workbook()
    sheet = book.create_sheet("Localized_Policies")

    with open(treef,'r') as tjaf:
        tjafile=tjaf.read()
        treef=json.loads(tjafile)
    

    rows=1
    sheet.cell(row=rows, column=1).value="Localized Policy"
    sheet.cell(row=rows, column=2).value="Device Template"
    sheet.cell(row=rows, column=3).value="Devices"
    sheet.cell(row=rows, column=4).value="Data/Control GUI policies"
   
    
    
    rows+=1
    for lp in treef["children"]:
        #sheet.cell(row=rows, column=1).value=lp["name"]
        s = ','.join(str(x) for x in lp["lptypes"])
        #sheet.cell(row=rows, column=4).value=s

        if len(lp["children"])>0:
            for dt in lp["children"]:
                sheet.cell(row=rows, column=2).value=dt["name"]
                if len(dt["children"])>0:
                    for device in dt["children"]:
                        sheet.cell(row=rows, column=1).value=lp["name"]
                        sheet.cell(row=rows, column=2).value=dt["name"]
                        sheet.cell(row=rows, column=3).value=device["name"]  
                        sheet.cell(row=rows, column=4).value=s

                        rows+=1
                else:
                    sheet.cell(row=rows, column=1).value=lp["name"]
                    sheet.cell(row=rows, column=2).value=dt["name"]
                    sheet.cell(row=rows, column=4).value=s
                    rows+=1

        else:
            sheet.cell(row=rows, column=1).value=lp["name"]
            sheet.cell(row=rows, column=4).value=s
            rows+=1 
            
       
                
    sheet.auto_filter.ref=sheet.dimensions
    book.save(excelf) 
    print("Spreadsheet file was created")


if __name__ == "__main__":

   #base_url="https://url/"

    

    vmanage_url=input("Please enter vManager url without https:// ")
    base_url=f"https://{vmanage_url}/"
    user=input("username:")
    password=getpass.getpass()


    dtjfile="device_temp.json"
    dtlpjfile="device_template_lp.json"
    lpjfile="locpol.json"
    lptreedata="lptreeData.json"
    excelf="SD-WAN Localized_Policies.xlsx"
    htmlfile="SD-WAN_Localized_Policies.html"


    jsessionid=get_jsessionid(base_url,user,password)
    token=get_token(jsessionid,base_url)
    header=setheader(jsessionid,token)
    getlp(header,base_url,lpjfile)
    getdt(header,base_url,dtjfile)
    getlpfromdt(header,base_url,dtjfile,dtlpjfile)
    get_lp_types(header,base_url,lpjfile)
    get_lp_config(header,base_url,lpjfile)
    create_relationship(lpjfile,dtjfile,dtlpjfile,header,base_url)
    buildtreemap(lpjfile,lptreedata)
    convert_excel(lptreedata,excelf)
    dtq,dq=count_dt_dev(lptreedata)
    create_html(lptreedata,htmlfile,dtq,dq)
    logouts(header,base_url)